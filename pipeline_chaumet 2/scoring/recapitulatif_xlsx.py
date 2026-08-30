"""Met en forme le récapitulatif des tentatives dans un classeur lisible.

Deux feuilles : « Tentatives » (une ligne par extraction demandée, filtrable) et
« Synthèse » (agrégats calculés par formules, donc recalculés si l'on filtre ou
corrige la première feuille).

Usage :
    python scoring/recapitulatif.py        # produit le CSV
    python scoring/recapitulatif_xlsx.py   # produit le classeur
"""
from __future__ import annotations

import csv
import re
import shutil
import zipfile
import sys
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "pipeline"))
import config  # noqa: E402

PARIS = ZoneInfo("Europe/Paris")
DOSSIER = config.BASE_DIR / "verification_croisee"

POLICE = "Arial"
TITRE = Font(name=POLICE, size=11, bold=True, color="FFFFFF")
NORMAL = Font(name=POLICE, size=10)
GRAS = Font(name=POLICE, size=10, bold=True)
ENTETE_FOND = PatternFill("solid", fgColor="404040")
ECHEC_FOND = PatternFill("solid", fgColor="FBE4E4")
MANQUE_FOND = PatternFill("solid", fgColor="FFF4D6")
FILET = Border(bottom=Side(style="thin", color="BFBFBF"))

COLONNES = [
    ("Cote d'archive", "cote_archive", 24),
    ("Modèle", "modele", 12),
    ("Série", "serie", 11),
    ("Exécution", "execution", 10),
    ("Horodatage (Paris)", "horodatage", 19),
    ("Durée (s)", "duree_s", 10),
    ("Issue", "issue", 10),
    ("Réessais", "reessais", 9),
    ("Cause de l'échec", "cause", 38),
    ("Message brut", "message", 70),
]


def lire() -> list[dict]:
    f = DOSSIER / "recapitulatif_fiches.csv"
    if not f.exists():
        raise SystemExit("Lancer d'abord : python scoring/recapitulatif.py")
    return list(csv.DictReader(f.open(encoding="utf-8-sig")))


def horodatage_local(iso: str) -> datetime | str:
    """L'horodatage d'une exécution reconstituée après coup n'existe pas ; on
    laisse alors la mention telle quelle plutôt que d'inventer une date."""
    if not iso:
        return ""
    try:
        return (datetime.fromisoformat(iso).replace(tzinfo=timezone.utc)
                .astimezone(PARIS).replace(tzinfo=None))
    except ValueError:
        return iso


def feuille_tentatives(wb: Workbook, lignes: list[dict]) -> int:
    ws = wb.active
    ws.title = "Tentatives"

    for i, (libelle, _, largeur) in enumerate(COLONNES, 1):
        c = ws.cell(1, i, libelle)
        c.font, c.fill = TITRE, ENTETE_FOND
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = largeur
    ws.row_dimensions[1].height = 30

    for r, l in enumerate(lignes, 2):
        echec = l["issue"] == "échec"
        for i, (_, cle, _) in enumerate(COLONNES, 1):
            v = l[cle]
            if cle == "horodatage":
                v = horodatage_local(v)
            elif cle in ("duree_s", "execution", "reessais"):
                v = float(v) if v else None
            c = ws.cell(r, i, v)
            c.font, c.border = NORMAL, FILET
            if cle == "horodatage":
                c.number_format = "yyyy-mm-dd hh:mm" if isinstance(v, datetime) else "@"
            elif cle == "duree_s":
                c.number_format = "0.0"
                if v is None and not echec:
                    c.fill = MANQUE_FOND
            elif cle in ("execution", "reessais"):
                c.number_format = "0"
            if cle in ("cote_archive", "modele", "serie", "issue"):
                c.alignment = Alignment(horizontal="left")
            if echec:
                c.fill = ECHEC_FOND
        ws.cell(r, 9).alignment = Alignment(wrap_text=False)

    fin = len(lignes) + 1
    ws.auto_filter.ref = f"A1:J{fin}"
    ws.freeze_panes = "A2"
    return fin


def feuille_synthese(wb: Workbook, fin: int, causes: list[str]) -> None:
    ws = wb.create_sheet("Synthèse")
    ws.column_dimensions["A"].width = 40
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 13

    def titre(r: int, texte: str) -> None:
        c = ws.cell(r, 1, texte)
        c.font = Font(name=POLICE, size=11, bold=True)

    def entete(r: int, libelles: list[str]) -> None:
        for i, lib in enumerate(libelles, 1):
            c = ws.cell(r, i, lib)
            c.font, c.fill = TITRE, ENTETE_FOND
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    T = f"Tentatives!$G$2:$G${fin}"      # issue
    M = f"Tentatives!$B$2:$B${fin}"      # modèle
    S = f"Tentatives!$C$2:$C${fin}"      # série
    D = f"Tentatives!$F$2:$F${fin}"      # durée
    C = f"Tentatives!$I$2:$I${fin}"      # cause

    titre(1, "Fiabilité d'exécution, par modèle et par série")
    ws.cell(1, 1).font = Font(name=POLICE, size=12, bold=True)
    entete(3, ["Modèle / série", "Tentatives", "Réussites", "Échecs",
               "Taux de réussite", "Durée moy. (s)", "Durée cumulée (h)"])

    r = 4
    for modele in ("Flash 3", "Pro 3.1"):
        for serie in ("référence", "rep_1", "rep_2", "TOTAL"):
            if serie == "TOTAL":
                cr = [(M, modele)]
                ws.cell(r, 1, f"{modele} — ensemble").font = GRAS
            else:
                cr = [(M, modele), (S, serie)]
                ws.cell(r, 1, f"{modele} — {serie}").font = NORMAL
            crit = "".join(f",{p},\"{v}\"" for p, v in cr)
            ws.cell(r, 2, f"=COUNTIFS({crit[1:]})")
            ws.cell(r, 3, f"=COUNTIFS({T},\"réussite\"{crit})")
            ws.cell(r, 4, f"=COUNTIFS({T},\"échec\"{crit})")
            ws.cell(r, 5, f"=IF(B{r}=0,\"\",C{r}/B{r})")
            ws.cell(r, 6, f"=IFERROR(AVERAGEIFS({D},{T},\"réussite\"{crit}),\"\")")
            ws.cell(r, 7, f"=SUMIFS({D},{T},\"réussite\"{crit})/3600")
            for i in range(2, 8):
                c = ws.cell(r, i)
                c.font = GRAS if serie == "TOTAL" else NORMAL
                c.number_format = {5: "0.0%", 6: "0.0", 7: "0.00"}.get(i, "0")
            for i in range(1, 8):
                ws.cell(r, i).border = FILET
            r += 1

    r += 1
    titre(r, "Causes d'échec")
    r += 1
    entete(r, ["Cause", "Flash 3", "Pro 3.1", "Total"])
    r += 1
    for cause in causes:
        ws.cell(r, 1, cause).font = NORMAL
        ws.cell(r, 2, f"=COUNTIFS({C},A{r},{M},\"Flash 3\")").font = NORMAL
        ws.cell(r, 3, f"=COUNTIFS({C},A{r},{M},\"Pro 3.1\")").font = NORMAL
        ws.cell(r, 4, f"=SUM(B{r}:C{r})").font = GRAS
        for i in range(1, 5):
            ws.cell(r, i).border = FILET
        r += 1
    ws.cell(r, 1, "Total").font = GRAS
    for i in (2, 3, 4):
        c = ws.cell(r, i, f"=SUM({get_column_letter(i)}{r-len(causes)}:{get_column_letter(i)}{r-1})")
        c.font = GRAS

    r += 2
    for note in [
        "Lecture du tableau",
        "• Une ligne de la feuille « Tentatives » = une extraction demandée à l'API, "
        "et non une fiche : une cote reprise après échec y figure autant de fois qu'elle a été soumise.",
        "• « Exécution » distingue les lancements successifs d'une même série "
        "(reprises après échec, option --reprendre).",
        "• L'horodatage est celui du début de l'exécution, seul instant consigné par les manifestes : "
        "il situe la série, pas la fiche.",
        "• « Réessais » compte les tentatives internes absorbées par la temporisation exponentielle. "
        "Cette colonne n'est renseignée que si scoring/journal.py a été exécuté sur l'export du terminal ; "
        "les manifestes ne conservent pas cette information.",
        "• Les 40 durées manquantes (fond ocre, Pro 3.1 rep_2) correspondent aux extractions récupérées "
        "après une interruption survenue avant l'écriture du manifeste : les fichiers de sortie existent, "
        "leur chronométrage est perdu. La moyenne de cette série ne porte donc que sur 10 fiches, "
        "et la durée cumulée de Pro 3.1 est sous-évaluée d'autant.",
        "• L'exécution lancée sur gemini-3-pro-preview, modèle retiré par Google en cours de projet, "
        "n'a produit aucun manifeste et ne figure pas ici : ses 19 tentatives ont toutes échoué en 404.",
        "• Les causes d'échec suivent la taxonomie de scoring/journal.py ; « corrigé » signale un défaut "
        "du pipeline depuis réparé, qui ne se reproduirait pas à l'identique.",
    ]:
        c = ws.cell(r, 1, note)
        c.font = GRAS if not note.startswith("•") else Font(name=POLICE, size=9)
        c.alignment = Alignment(wrap_text=False)
        r += 1


def injecter_valeurs(sortie: Path, lignes: list[dict], causes: list[str]) -> int:
    """Dépose dans le XML la valeur de chaque formule de la feuille « Synthèse ».

    openpyxl écrit les formules sans valeur mémorisée : tout lecteur qui ne
    recalcule pas — aperçu du Finder, pandas, load_workbook(data_only=True) —
    voit des cellules vides. On calcule donc les mêmes agrégats en Python et on
    les inscrit à côté des formules, qu'Excel remplacera au premier recalcul.
    Les deux chemins de calcul étant indépendants, leur concordance vaut
    vérification.
    """
    def sel(**kw):
        return [l for l in lignes if all(l[k] == v for k, v in kw.items())]

    valeurs: dict[str, float | int] = {}
    r = 4
    for modele in ("Flash 3", "Pro 3.1"):
        for serie in ("référence", "rep_1", "rep_2", "TOTAL"):
            g = sel(modele=modele) if serie == "TOTAL" else sel(modele=modele, serie=serie)
            ok = [l for l in g if l["issue"] == "réussite"]
            d = [float(l["duree_s"]) for l in ok if l["duree_s"]]
            valeurs |= {f"B{r}": len(g), f"C{r}": len(ok), f"D{r}": len(g) - len(ok),
                        f"E{r}": len(ok) / len(g), f"F{r}": sum(d) / len(d),
                        f"G{r}": sum(d) / 3600}
            r += 1

    # Le bloc des causes est repéré dans le classeur : recalculer son décalage
    # le rendrait faux à la moindre retouche de la mise en page.
    ws = load_workbook(sortie)["Synthèse"]
    r = premiere = next(c.row for c in ws["A"] if c.value == causes[0])
    for cause in causes:
        f = len([l for l in lignes if l["cause"] == cause and l["modele"] == "Flash 3"])
        p = len([l for l in lignes if l["cause"] == cause and l["modele"] == "Pro 3.1"])
        valeurs |= {f"B{r}": f, f"C{r}": p, f"D{r}": f + p}
        r += 1
    for col in "BCD":
        valeurs[f"{col}{r}"] = sum(valeurs[f"{col}{i}"] for i in range(premiere, r))

    poses = 0

    def remplacer(m: "re.Match[str]") -> str:
        nonlocal poses
        corps, ref = m.group(0), m.group(1)
        if ref not in valeurs or "<f>" not in corps:
            return corps
        poses += 1
        v = valeurs[ref]
        texte = repr(round(v, 10)) if isinstance(v, float) else str(v)
        return re.sub(r"<v>\s*</v>", "", corps).replace("</f>", f"</f><v>{texte}</v>")

    temporaire = sortie.with_suffix(".tmp")
    with zipfile.ZipFile(sortie) as zin, \
         zipfile.ZipFile(temporaire, "w", zipfile.ZIP_DEFLATED) as zout:
        cible = f"xl/worksheets/sheet{wb_index(zin)}.xml"
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == cible:
                data = re.sub(r'<c r="([A-Z]+\d+)"[^>]*>.*?</c>', remplacer,
                              data.decode("utf-8"), flags=re.S).encode("utf-8")
            zout.writestr(item, data)
    shutil.move(temporaire, sortie)

    if poses != len(valeurs):
        raise SystemExit(f"{poses} valeurs posées sur {len(valeurs)} attendues : "
                         "la mise en page a changé, vérifier injecter_valeurs().")
    return poses


def wb_index(zin: zipfile.ZipFile) -> int:
    """Rang de la feuille « Synthèse » parmi les worksheets du classeur."""
    noms = re.findall(r'<sheet [^>]*name="([^"]+)"', zin.read("xl/workbook.xml").decode())
    return noms.index("Synthèse") + 1


def main() -> None:
    lignes = lire()
    causes = sorted({l["cause"] for l in lignes if l["cause"]})
    wb = Workbook()
    fin = feuille_tentatives(wb, lignes)
    feuille_synthese(wb, fin, causes)
    sortie = DOSSIER / "recapitulatif_tentatives.xlsx"
    wb.save(sortie)
    n = injecter_valeurs(sortie, lignes, causes)
    print(f"Écrit : {sortie}\n  {len(lignes)} tentatives, {n} agrégats vérifiés")


if __name__ == "__main__":
    main()
